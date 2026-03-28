from creditriskengine.rwa.irb.formulas import irb_risk_weight
from creditriskengine.rwa.standardized.credit_risk_sa import assign_sa_risk_weight
from creditriskengine.models.ead.ead_model import calculate_ead, get_supervisory_ccf
from creditriskengine.core.types import SAExposureClass, CreditQualityStep, Jurisdiction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── HELPERS ────────────────────────────────────────────────────────────────

def make_loan(name, irb_class, sa_class, cqs, pd, lgd, drawn, undrawn=0,
              facility="term_loan", maturity=2.5, ltv=None, domestic=False,
              sector="Corporate", country="India", rating="BBB"):
    ccf = get_supervisory_ccf(facility) if undrawn > 0 else 1.0
    ead = calculate_ead(drawn, undrawn, ccf)
    return {
        "name": name, "sector": sector, "country": country,
        "instrument_type": "Loan/Bond",
        "irb_class": irb_class, "sa_class": sa_class, "cqs": cqs,
        "rating": rating, "pd": pd, "lgd": lgd,
        "drawn": drawn, "undrawn": undrawn, "ccf": ccf, "ead": ead,
        "maturity": maturity, "ltv": ltv if ltv else "-",
        "domestic": domestic, "is_derivative": False,
        "notional": drawn + undrawn, "mtm": "-", "add_on_pct": "-",
        "ead_note": "Drawn + CCF x Undrawn" if undrawn > 0 else "Fully drawn",
    }

def make_derivative(name, asset_class, sa_class, cqs, pd, lgd,
                    mtm, notional, add_on_pct, maturity=2.5,
                    sector="Derivatives", country="India",
                    deriv_type="IRS", rating="A"):
    pfe = notional * add_on_pct
    ead = max(mtm, 0) + pfe
    return {
        "name": name, "sector": sector, "country": country,
        "instrument_type": deriv_type,
        "irb_class": asset_class, "sa_class": sa_class, "cqs": cqs,
        "rating": rating, "pd": pd, "lgd": lgd,
        "drawn": "-", "undrawn": "-", "ccf": "-", "ead": ead,
        "maturity": maturity, "ltv": "-",
        "domestic": False, "is_derivative": True,
        "notional": notional, "mtm": mtm, "add_on_pct": add_on_pct,
        "ead_note": f"max(MTM,0) + Notional x {add_on_pct:.1%} add-on",
    }

# ── PORTFOLIO ──────────────────────────────────────────────────────────────
portfolio = []

# SOVEREIGNS
portfolio += [
    make_loan("Govt of India - INR",      "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_3, 0.005, 0.00, 50_000_000, domestic=True,  sector="Sovereign", country="India",    rating="BBB"),
    make_loan("Govt of India - USD",      "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_3, 0.008, 0.45, 20_000_000, domestic=False, sector="Sovereign", country="India",    rating="BBB"),
    make_loan("US Treasury 2Y",           "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.001, 0.00, 30_000_000, domestic=False, sector="Sovereign", country="USA",      rating="AAA"),
    make_loan("US Treasury 10Y",          "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.001, 0.00, 20_000_000, domestic=False, sector="Sovereign", country="USA",      rating="AAA"),
    make_loan("UK Gilt 5Y",               "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.001, 0.00, 15_000_000, domestic=False, sector="Sovereign", country="UK",       rating="AA"),
    make_loan("German Bund 10Y",          "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.001, 0.00, 10_000_000, domestic=False, sector="Sovereign", country="Germany",  rating="AAA"),
    make_loan("French OAT",               "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.002, 0.00,  8_000_000, domestic=False, sector="Sovereign", country="France",   rating="AA"),
    make_loan("Japan JGB",                "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.001, 0.00, 12_000_000, domestic=False, sector="Sovereign", country="Japan",    rating="A"),
    make_loan("Singapore SGS",            "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_1, 0.001, 0.00,  6_000_000, domestic=False, sector="Sovereign", country="Singapore",rating="AAA"),
    make_loan("Brazil Sovereign",         "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_4, 0.030, 0.45,  5_000_000, domestic=False, sector="Sovereign", country="Brazil",   rating="BB"),
    make_loan("South Africa Sovereign",   "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_4, 0.035, 0.45,  4_000_000, domestic=False, sector="Sovereign", country="S.Africa", rating="BB"),
    make_loan("Indonesia Sovereign",      "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_3, 0.020, 0.45,  5_000_000, domestic=False, sector="Sovereign", country="Indonesia",rating="BBB"),
    make_loan("Mexico Sovereign",         "sovereign", SAExposureClass.SOVEREIGN, CreditQualityStep.CQS_3, 0.022, 0.45,  3_000_000, domestic=False, sector="Sovereign", country="Mexico",   rating="BBB"),
]

# BANKS
portfolio += [
    make_loan("HDFC Bank",           "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.008, 0.40, 10_000_000, sector="Banking", country="India",      rating="A"),
    make_loan("SBI",                 "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.012, 0.45,  8_000_000, sector="Banking", country="India",      rating="A"),
    make_loan("ICICI Bank",          "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  7_000_000, sector="Banking", country="India",      rating="A"),
    make_loan("Axis Bank",           "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.011, 0.42,  5_000_000, sector="Banking", country="India",      rating="A"),
    make_loan("Kotak Mahindra Bank", "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.009, 0.40,  6_000_000, sector="Banking", country="India",      rating="A"),
    make_loan("JP Morgan",           "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35, 12_000_000, sector="Banking", country="USA",        rating="AA"),
    make_loan("Goldman Sachs",       "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.006, 0.35,  8_000_000, sector="Banking", country="USA",        rating="A"),
    make_loan("Deutsche Bank",       "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.012, 0.40,  6_000_000, sector="Banking", country="Germany",    rating="A"),
    make_loan("BNP Paribas",         "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.007, 0.38,  7_000_000, sector="Banking", country="France",     rating="A"),
    make_loan("Barclays",            "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  8_000_000, sector="Banking", country="UK",         rating="A"),
    make_loan("Standard Chartered",  "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.011, 0.40,  5_000_000, sector="Banking", country="UK",         rating="A"),
    make_loan("MUFG",                "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.006, 0.35,  7_000_000, sector="Banking", country="Japan",      rating="A"),
    make_loan("DBS Bank",            "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35,  5_000_000, sector="Banking", country="Singapore",  rating="AA"),
    make_loan("UBS",                 "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.006, 0.35,  6_000_000, sector="Banking", country="Switzerland",rating="A"),
    make_loan("HSBC",                "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.007, 0.37,  9_000_000, sector="Banking", country="UK",         rating="A"),
]

# CORPORATES
portfolio += [
    make_loan("Tata Steel",          "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.020, 0.45,  5_000_000, 2_000_000, "committed_other", 3.0, sector="Manufacturing",  country="India",    rating="BBB"),
    make_loan("Reliance Industries", "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.012, 0.40,  8_000_000, 3_000_000, "committed_other", 4.0, sector="Energy",         country="India",    rating="A"),
    make_loan("Infosys",             "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  6_000_000, 0,         "term_loan",       3.0, sector="Technology",     country="India",    rating="A"),
    make_loan("ONGC",                "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.018, 0.42,  7_000_000, 2_500_000, "committed_other", 3.5, sector="Energy",         country="India",    rating="BBB"),
    make_loan("Adani Ports",         "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.022, 0.45,  4_000_000, 1_000_000, "committed_other", 4.0, sector="Infrastructure", country="India",    rating="BBB"),
    make_loan("Wipro",               "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  3_500_000, 0,         "term_loan",       2.5, sector="Technology",     country="India",    rating="A"),
    make_loan("L&T",                 "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.015, 0.42,  5_000_000, 2_000_000, "committed_other", 3.0, sector="Infrastructure", country="India",    rating="A"),
    make_loan("ITC Ltd",             "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  3_000_000, 0,         "term_loan",       2.5, sector="FMCG",           country="India",    rating="A"),
    make_loan("Bharat Petroleum",    "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.018, 0.42,  4_000_000, 1_000_000, "committed_other", 3.0, sector="Energy",         country="India",    rating="BBB"),
    make_loan("Mahindra",            "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.020, 0.43,  3_500_000,   500_000, "committed_other", 3.0, sector="Auto",           country="India",    rating="BBB"),
    make_loan("JSW Steel",           "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.025, 0.46,  4_000_000, 1_000_000, "committed_other", 3.5, sector="Manufacturing",  country="India",    rating="BBB"),
    make_loan("UltraTech Cement",    "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.018, 0.42,  3_000_000, 0,         "term_loan",       3.0, sector="Manufacturing",  country="India",    rating="BBB"),
    make_loan("Bajaj Finance",       "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.015, 0.40,  4_000_000, 1_000_000, "committed_other", 2.5, sector="NBFC",           country="India",    rating="A"),
    make_loan("HCL Technologies",    "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  3_500_000, 0,         "term_loan",       3.0, sector="Technology",     country="India",    rating="A"),
    make_loan("Sun Pharma",          "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.012, 0.40,  3_000_000, 0,         "term_loan",       2.5, sector="Healthcare",     country="India",    rating="A"),
    make_loan("Asian Paints",        "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  2_500_000, 0,         "term_loan",       2.5, sector="Manufacturing",  country="India",    rating="A"),
    make_loan("Maruti Suzuki",       "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.012, 0.40,  4_000_000, 0,         "term_loan",       3.0, sector="Auto",           country="India",    rating="A"),
    make_loan("Apple Inc",           "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_1, 0.005, 0.35, 10_000_000, 0,         "term_loan",       3.0, sector="Technology",     country="USA",      rating="AA"),
    make_loan("Amazon",              "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_1, 0.006, 0.35,  8_000_000, 2_000_000, "committed_other", 3.0, sector="Technology",     country="USA",      rating="AA"),
    make_loan("Microsoft",           "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_1, 0.004, 0.33,  9_000_000, 0,         "term_loan",       3.0, sector="Technology",     country="USA",      rating="AAA"),
    make_loan("ExxonMobil",          "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  6_000_000, 0,         "term_loan",       4.0, sector="Energy",         country="USA",      rating="A"),
    make_loan("Shell",               "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.012, 0.40,  6_000_000, 1_500_000, "committed_other", 4.0, sector="Energy",         country="UK",       rating="A"),
    make_loan("BP",                  "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.013, 0.40,  5_000_000, 0,         "term_loan",       4.0, sector="Energy",         country="UK",       rating="A"),
    make_loan("Volkswagen",          "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_3, 0.020, 0.43,  5_000_000, 0,         "term_loan",       3.0, sector="Auto",           country="Germany",  rating="BBB"),
    make_loan("Siemens",             "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  4_000_000, 0,         "term_loan",       3.0, sector="Manufacturing",  country="Germany",  rating="A"),
    make_loan("LVMH",                "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.008, 0.37,  5_000_000, 0,         "term_loan",       3.0, sector="FMCG",           country="France",   rating="A"),
    make_loan("Toyota",              "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.010, 0.38,  6_000_000, 0,         "term_loan",       3.0, sector="Auto",           country="Japan",    rating="A"),
    make_loan("Samsung",             "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_2, 0.009, 0.37,  5_000_000, 0,         "term_loan",       3.0, sector="Technology",     country="S.Korea",  rating="A"),
    make_loan("Vale SA",             "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_4, 0.035, 0.50,  3_000_000, 0,         "term_loan",       3.0, sector="Mining",         country="Brazil",   rating="BB"),
    make_loan("Petrobras",           "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_4, 0.038, 0.52,  2_500_000, 0,         "term_loan",       3.0, sector="Energy",         country="Brazil",   rating="BB"),
    make_loan("Eskom",               "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_5, 0.060, 0.55,  2_000_000, 0,         "term_loan",       2.0, sector="Utilities",      country="S.Africa", rating="B"),
    make_loan("Sasol",               "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_4, 0.040, 0.50,  1_500_000, 0,         "term_loan",       2.0, sector="Energy",         country="S.Africa", rating="BB"),
    make_loan("Telkom SA",           "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_4, 0.038, 0.50,  1_200_000, 0,         "term_loan",       2.0, sector="Telecom",        country="S.Africa", rating="BB"),
    make_loan("Grab Holdings",       "corporate", SAExposureClass.CORPORATE, CreditQualityStep.CQS_4, 0.042, 0.52,  2_000_000, 0,         "term_loan",       3.0, sector="Technology",     country="Singapore",rating="BB"),
]

# SME CORPORATES
portfolio += [
    make_loan("SME - Auto Parts",    "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.035, 0.50, 800_000, 200_000, "committed_other", 2.0, sector="Manufacturing",  country="India", rating="Unrated"),
    make_loan("SME - Textile Mill",  "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.040, 0.52, 600_000, 150_000, "committed_other", 2.0, sector="Manufacturing",  country="India", rating="Unrated"),
    make_loan("SME - Pharma Unit",   "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.030, 0.48, 900_000, 100_000, "committed_other", 2.0, sector="Healthcare",     country="India", rating="Unrated"),
    make_loan("SME - IT Services",   "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.025, 0.45, 500_000, 0,       "term_loan",       2.0, sector="Technology",     country="India", rating="Unrated"),
    make_loan("SME - Food Proc.",    "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.038, 0.50, 700_000, 200_000, "committed_other", 2.0, sector="FMCG",           country="India", rating="Unrated"),
    make_loan("SME - Construction",  "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.045, 0.55, 600_000, 200_000, "committed_other", 2.0, sector="Infrastructure", country="India", rating="Unrated"),
    make_loan("SME - Logistics",     "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.032, 0.48, 750_000, 100_000, "committed_other", 2.0, sector="Logistics",      country="India", rating="Unrated"),
    make_loan("SME - Healthcare",    "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.028, 0.46, 850_000, 0,       "term_loan",       2.0, sector="Healthcare",     country="India", rating="Unrated"),
    make_loan("SME - Agri Inputs",   "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.036, 0.50, 450_000, 100_000, "committed_other", 2.0, sector="Agriculture",    country="India", rating="Unrated"),
    make_loan("SME - Retail Chain",  "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.042, 0.52, 550_000, 0,       "term_loan",       2.0, sector="Retail",         country="India", rating="Unrated"),
    make_loan("SME - Steel Fab",     "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.040, 0.52, 650_000, 150_000, "committed_other", 2.0, sector="Manufacturing",  country="India", rating="Unrated"),
    make_loan("SME - Chemicals",     "corporate", SAExposureClass.CORPORATE_SME, CreditQualityStep.UNRATED, 0.033, 0.49, 700_000, 0,       "term_loan",       2.0, sector="Manufacturing",  country="India", rating="Unrated"),
]

# RESIDENTIAL MORTGAGES
portfolio += [
    make_loan("Home Loan - Mumbai A",    "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.020, 0.15, 3_000_000, ltv=0.65, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Mumbai B",    "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.025, 0.18, 2_500_000, ltv=0.72, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Mumbai C",    "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.022, 0.16, 2_000_000, ltv=0.68, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Delhi A",     "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.022, 0.16, 2_000_000, ltv=0.70, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Delhi B",     "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.028, 0.20, 1_800_000, ltv=0.78, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Delhi C",     "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.030, 0.22, 1_500_000, ltv=0.82, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Bangalore A", "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.018, 0.14, 2_200_000, ltv=0.68, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Bangalore B", "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.021, 0.16, 1_800_000, ltv=0.72, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Chennai A",   "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.023, 0.17, 1_500_000, ltv=0.75, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Chennai B",   "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.026, 0.19, 1_200_000, ltv=0.79, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Pune A",      "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.021, 0.15, 1_200_000, ltv=0.71, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Pune B",      "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.024, 0.17, 1_000_000, ltv=0.74, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Hyderabad A", "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.024, 0.18, 1_600_000, ltv=0.73, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Hyderabad B", "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.027, 0.20, 1_300_000, ltv=0.77, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Ahmedabad",   "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.022, 0.16, 1_100_000, ltv=0.70, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - Kolkata",     "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.025, 0.18, 1_000_000, ltv=0.73, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - High LTV A",  "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.035, 0.25, 2_000_000, ltv=0.88, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - High LTV B",  "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.040, 0.28, 1_500_000, ltv=0.92, sector="Retail", country="India", rating="Unrated"),
    make_loan("Home Loan - High LTV C",  "residential_mortgage", SAExposureClass.RESIDENTIAL_MORTGAGE, CreditQualityStep.UNRATED, 0.045, 0.30, 1_200_000, ltv=0.95, sector="Retail", country="India", rating="Unrated"),
]

# RETAIL
portfolio += [
    make_loan("Credit Card Pool A",  "qrre",         SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.050, 0.80,   500_000, 200_000, "committed_unconditionally_cancellable", 1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Credit Card Pool B",  "qrre",         SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.065, 0.80,   400_000, 150_000, "committed_unconditionally_cancellable", 1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Credit Card Pool C",  "qrre",         SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.045, 0.75,   600_000, 250_000, "committed_unconditionally_cancellable", 1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Credit Card Pool D",  "qrre",         SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.055, 0.78,   350_000, 100_000, "committed_unconditionally_cancellable", 1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Personal Loan A",     "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.040, 0.60,   300_000, 0,       "term_loan",                            2.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Personal Loan B",     "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.045, 0.62,   250_000, 0,       "term_loan",                            2.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Personal Loan C",     "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.055, 0.65,   200_000, 0,       "term_loan",                            1.5, sector="Retail", country="India", rating="Unrated"),
    make_loan("Personal Loan D",     "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.048, 0.63,   280_000, 0,       "term_loan",                            2.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Auto Loan Pool A",    "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.030, 0.40,   800_000, 0,       "term_loan",                            3.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Auto Loan Pool B",    "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.035, 0.42,   600_000, 0,       "term_loan",                            3.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Auto Loan Pool C",    "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.038, 0.44,   500_000, 0,       "term_loan",                            3.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Gold Loan Pool A",    "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.020, 0.25, 1_000_000, 0,       "term_loan",                            1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Gold Loan Pool B",    "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.022, 0.26,   800_000, 0,       "term_loan",                            1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Microfinance Pool A", "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.060, 0.70,   500_000, 0,       "term_loan",                            1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Microfinance Pool B", "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.065, 0.72,   400_000, 0,       "term_loan",                            1.0, sector="Retail", country="India", rating="Unrated"),
    make_loan("Education Loan Pool", "other_retail", SAExposureClass.RETAIL, CreditQualityStep.UNRATED, 0.035, 0.50,   600_000, 0,       "term_loan",                            5.0, sector="Retail", country="India", rating="Unrated"),
]

# DERIVATIVES
portfolio += [
    make_derivative("IRS - Reliance 5Y",       "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  150_000,  50_000_000, 0.005, maturity=5.0,  sector="Derivatives", country="India",    deriv_type="Interest Rate Swap",      rating="A"),
    make_derivative("IRS - SBI 3Y",            "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.012, 0.45,  -80_000,  30_000_000, 0.005, maturity=3.0,  sector="Derivatives", country="India",    deriv_type="Interest Rate Swap",      rating="A"),
    make_derivative("IRS - HDFC 7Y",           "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.009, 0.40,  220_000,  40_000_000, 0.008, maturity=7.0,  sector="Derivatives", country="India",    deriv_type="Interest Rate Swap",      rating="A"),
    make_derivative("IRS - JP Morgan 10Y",     "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35,  500_000, 100_000_000, 0.015, maturity=10.0, sector="Derivatives", country="USA",      deriv_type="Interest Rate Swap",      rating="AA"),
    make_derivative("IRS - Barclays 5Y",       "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  180_000,  60_000_000, 0.005, maturity=5.0,  sector="Derivatives", country="UK",       deriv_type="Interest Rate Swap",      rating="A"),
    make_derivative("CCS - USD/INR 2Y",        "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  200_000,  20_000_000, 0.015, maturity=2.0,  sector="Derivatives", country="India",    deriv_type="Cross Currency Swap",     rating="A"),
    make_derivative("CCS - EUR/INR 3Y",        "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.011, 0.40,  120_000,  15_000_000, 0.015, maturity=3.0,  sector="Derivatives", country="India",    deriv_type="Cross Currency Swap",     rating="A"),
    make_derivative("CCS - JPY/USD 5Y",        "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.006, 0.35,  350_000,  25_000_000, 0.020, maturity=5.0,  sector="Derivatives", country="Japan",    deriv_type="Cross Currency Swap",     rating="AA"),
    make_derivative("FX Forward - EUR/USD 1Y", "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35,   50_000,  10_000_000, 0.010, maturity=1.0,  sector="Derivatives", country="Germany",  deriv_type="FX Forward",              rating="AA"),
    make_derivative("FX Forward - GBP/USD 1Y", "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35,   30_000,   8_000_000, 0.010, maturity=1.0,  sector="Derivatives", country="UK",       deriv_type="FX Forward",              rating="AA"),
    make_derivative("FX Option - USD/INR",     "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,   80_000,  12_000_000, 0.010, maturity=0.5,  sector="Derivatives", country="India",    deriv_type="FX Option",               rating="A"),
    make_derivative("FX Option - EUR/INR",     "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,   60_000,   8_000_000, 0.010, maturity=0.5,  sector="Derivatives", country="India",    deriv_type="FX Option",               rating="A"),
    make_derivative("Equity TRS - Nifty",      "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  300_000,  15_000_000, 0.060, maturity=1.0,  sector="Derivatives", country="India",    deriv_type="Equity Total Return Swap",rating="A"),
    make_derivative("Equity Option - Sensex",  "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.012, 0.42,  150_000,  10_000_000, 0.060, maturity=0.5,  sector="Derivatives", country="India",    deriv_type="Equity Option",           rating="A"),
    make_derivative("Equity Fwd - S&P500",     "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35,  250_000,  20_000_000, 0.060, maturity=1.0,  sector="Derivatives", country="USA",      deriv_type="Equity Forward",          rating="AA"),
    make_derivative("Equity Swap - FTSE",      "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.40,  180_000,  12_000_000, 0.060, maturity=1.0,  sector="Derivatives", country="UK",       deriv_type="Equity Total Return Swap",rating="A"),
    make_derivative("CDS - Tata Steel",        "bank", SAExposureClass.BANK, CreditQualityStep.CQS_3, 0.020, 0.45,  100_000,  25_000_000, 0.050, maturity=3.0,  sector="Derivatives", country="India",    deriv_type="Credit Default Swap",     rating="BBB"),
    make_derivative("CDS - Reliance",          "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.012, 0.40,   80_000,  20_000_000, 0.050, maturity=5.0,  sector="Derivatives", country="India",    deriv_type="Credit Default Swap",     rating="A"),
    make_derivative("CDS - Brazil Sov",        "bank", SAExposureClass.BANK, CreditQualityStep.CQS_4, 0.030, 0.45,  200_000,  15_000_000, 0.050, maturity=5.0,  sector="Derivatives", country="Brazil",   deriv_type="Credit Default Swap",     rating="BB"),
    make_derivative("Commodity Swap - Oil",    "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.012, 0.40,  -20_000,   8_000_000, 0.150, maturity=2.0,  sector="Derivatives", country="India",    deriv_type="Commodity Swap",          rating="A"),
    make_derivative("Commodity Swap - Gold",   "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.010, 0.38,   90_000,   5_000_000, 0.150, maturity=1.0,  sector="Derivatives", country="India",    deriv_type="Commodity Swap",          rating="A"),
    make_derivative("Commodity Fwd - Copper",  "bank", SAExposureClass.BANK, CreditQualityStep.CQS_3, 0.018, 0.42,   40_000,   3_000_000, 0.150, maturity=0.5,  sector="Derivatives", country="India",    deriv_type="Commodity Forward",       rating="BBB"),
    make_derivative("Bond Fwd - G-Sec 10Y",    "bank", SAExposureClass.BANK, CreditQualityStep.CQS_2, 0.008, 0.38,  120_000,  30_000_000, 0.005, maturity=1.0,  sector="Derivatives", country="India",    deriv_type="Bond Forward",            rating="A"),
    make_derivative("Swaption - 5Y x 5Y",      "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.006, 0.35,  280_000,  50_000_000, 0.008, maturity=5.0,  sector="Derivatives", country="USA",      deriv_type="Swaption",                rating="AA"),
    make_derivative("Bond Fwd - UK Gilt",       "bank", SAExposureClass.BANK, CreditQualityStep.CQS_1, 0.005, 0.35,   90_000,  20_000_000, 0.005, maturity=1.0,  sector="Derivatives", country="UK",       deriv_type="Bond Forward",            rating="AA"),
]

# ── CALCULATIONS ───────────────────────────────────────────────────────────
scenarios = {"Baseline": 1.0, "Adverse": 1.5, "Severely Adverse": 2.5}
jurisdictions = [Jurisdiction.BCBS, Jurisdiction.EU, Jurisdiction.UK, Jurisdiction.INDIA]

def stressed_pd(pd, mult): return min(pd * mult, 1.0)

def get_sa_rw(loan, jur):
    kwargs = {"exposure_class": loan["sa_class"], "cqs": loan["cqs"], "jurisdiction": jur}
    if isinstance(loan["ltv"], float): kwargs["ltv"] = loan["ltv"]
    if loan["domestic"]: kwargs["is_domestic_own_currency"] = True
    if loan["sa_class"] == SAExposureClass.CORPORATE_SME: kwargs["is_sme"] = True
    return assign_sa_risk_weight(**kwargs)

for loan in portfolio:
    loan["irb_rw_baseline"]      = irb_risk_weight(pd=loan["pd"], lgd=loan["lgd"], asset_class=loan["irb_class"], maturity=loan["maturity"])
    loan["irb_rwa_baseline"]     = (loan["irb_rw_baseline"] / 100) * loan["ead"]
    loan["irb_capital_baseline"] = loan["irb_rwa_baseline"] * 0.08
    for scenario, mult in scenarios.items():
        spd = stressed_pd(loan["pd"], mult)
        rw  = irb_risk_weight(pd=spd, lgd=loan["lgd"], asset_class=loan["irb_class"], maturity=loan["maturity"])
        loan[f"stressed_pd_{scenario}"] = spd
        loan[f"irb_rw_{scenario}"]      = rw
        loan[f"irb_rwa_{scenario}"]     = (rw / 100) * loan["ead"]
    for jur in jurisdictions:
        rw = get_sa_rw(loan, jur)
        loan[f"sa_rw_{jur.name}"]  = rw
        loan[f"sa_rwa_{jur.name}"] = (rw / 100) * loan["ead"]

total_ead = sum(l["ead"] for l in portfolio)

# ── EXCEL EXPORT ───────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
YELLOW     = "FFF2CC"
GREEN      = "E2EFDA"

def hdr_font(size=10, bold=True, color=WHITE):   return Font(name="Arial", size=size, bold=bold, color=color)
def body_font(size=9, bold=False, color="000000"): return Font(name="Arial", size=size, bold=bold, color=color)
def fill(hex_color):  return PatternFill("solid", fgColor=hex_color)
def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def right():  return Alignment(horizontal="right",  vertical="center")

def apply_header_row(ws, row, cols, bg=DARK_BLUE, font_color=WHITE):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = hdr_font(color=font_color)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = border()

def apply_data_row(ws, row_num, values, bg=WHITE, number_formats=None, blue_cols=None):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.font = body_font(color="0000FF" if blue_cols and c in blue_cols else "000000")
        cell.fill = fill(bg)
        cell.alignment = right() if isinstance(val, (int, float)) else Alignment(vertical="center")
        cell.border = border()
        if number_formats and c <= len(number_formats) and number_formats[c-1]:
            cell.number_format = number_formats[c-1]

wb = Workbook()

# ── SHEET 1: PORTFOLIO DETAIL ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Portfolio Detail"
ws1.freeze_panes = "A3"
ws1.merge_cells("A1:AJ1")
ws1["A1"].value = "ICAAP STRESS TESTING & JURISDICTION RWA DASHBOARD — Portfolio Detail"
ws1["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws1["A1"].fill = fill(DARK_BLUE)
ws1["A1"].alignment = center()
ws1.row_dimensions[1].height = 28

headers = [
    "No.", "Exposure Name", "Sector", "Country", "Instrument Type",
    "IRB Asset Class", "Rating", "Maturity (Yrs)",
    "Drawn Amount", "Undrawn Commit.", "CCF", "Notional", "MTM Value", "Add-On %",
    "EAD", "EAD Calculation Note",
    "PD (Baseline)", "LGD", "LTV",
    "IRB Risk Weight %", "IRB RWA", "IRB Capital (8%)",
    "Adverse PD", "Adverse RW%", "Adverse RWA",
    "Sev.Adverse PD", "Sev.Adverse RW%", "Sev.Adverse RWA",
    "SA RW% - BCBS", "SA RWA - BCBS",
    "SA RW% - EU",   "SA RWA - EU",
    "SA RW% - UK",   "SA RWA - UK",
    "SA RW% - INDIA","SA RWA - INDIA",
]
apply_header_row(ws1, 2, headers)
ws1.row_dimensions[2].height = 40

col_widths = [5,28,14,12,22,20,8,10, 14,14,8,14,12,10, 14,30, 12,10,8, 14,14,12, 12,12,14, 14,14,14, 12,14,12,14,12,14,12,14]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

nf = [None,None,None,None,None,None,None,"0.0",
      '#,##0;(#,##0);"-"','#,##0;(#,##0);"-"','0.00;-;"-"',
      '#,##0;(#,##0);"-"','#,##0;(#,##0);"-"','0.0%;-;"-"',
      '#,##0',None,
      '0.000%','0.0%','0.0%;-;"-"',
      '0.00%','#,##0','#,##0',
      '0.000%','0.00%','#,##0',
      '0.000%','0.00%','#,##0',
      '0.00%','#,##0','0.00%','#,##0','0.00%','#,##0','0.00%','#,##0']

blue_input_cols = {9,10,12,13,14,17,18,19}

for i, loan in enumerate(portfolio):
    r   = i + 3
    bg  = LIGHT_GREY if i % 2 == 0 else WHITE
    ltv = loan["ltv"] if isinstance(loan["ltv"], float) else None
    row_vals = [
        i+1, loan["name"], loan["sector"], loan["country"], loan["instrument_type"],
        loan["irb_class"], loan["rating"], loan["maturity"],
        loan["drawn"] if not loan["is_derivative"] else None,
        loan["undrawn"] if not loan["is_derivative"] else None,
        loan["ccf"] if not loan["is_derivative"] else None,
        loan["notional"],
        loan["mtm"] if loan["is_derivative"] else None,
        loan["add_on_pct"] if loan["is_derivative"] else None,
        loan["ead"], loan["ead_note"],
        loan["pd"], loan["lgd"], ltv,
        loan["irb_rw_baseline"]/100, loan["irb_rwa_baseline"], loan["irb_capital_baseline"],
        loan["stressed_pd_Adverse"], loan["irb_rw_Adverse"]/100, loan["irb_rwa_Adverse"],
        loan["stressed_pd_Severely Adverse"], loan["irb_rw_Severely Adverse"]/100, loan["irb_rwa_Severely Adverse"],
        loan["sa_rw_BCBS"]/100,  loan["sa_rwa_BCBS"],
        loan["sa_rw_EU"]/100,    loan["sa_rwa_EU"],
        loan["sa_rw_UK"]/100,    loan["sa_rwa_UK"],
        loan["sa_rw_INDIA"]/100, loan["sa_rwa_INDIA"],
    ]
    apply_data_row(ws1, r, row_vals, bg=bg, number_formats=nf, blue_cols=blue_input_cols)

r_total = len(portfolio) + 3
ws1.cell(r_total, 1).value = "TOTAL"
ws1.cell(r_total, 1).font  = Font(name="Arial", size=9, bold=True, color=WHITE)
ws1.cell(r_total, 1).fill  = fill(MID_BLUE)
for col, val in {
    15: total_ead,
    21: sum(l["irb_rwa_baseline"] for l in portfolio),
    22: sum(l["irb_capital_baseline"] for l in portfolio),
    25: sum(l["irb_rwa_Adverse"] for l in portfolio),
    28: sum(l["irb_rwa_Severely Adverse"] for l in portfolio),
    30: sum(l["sa_rwa_BCBS"] for l in portfolio),
    32: sum(l["sa_rwa_EU"] for l in portfolio),
    34: sum(l["sa_rwa_UK"] for l in portfolio),
    36: sum(l["sa_rwa_INDIA"] for l in portfolio),
}.items():
    cell = ws1.cell(r_total, col, value=val)
    cell.font   = Font(name="Arial", size=9, bold=True, color=WHITE)
    cell.fill   = fill(MID_BLUE)
    cell.number_format = '#,##0'
    cell.alignment = right()
    cell.border = border()

# ── SHEET 2: IRB STRESS TEST ───────────────────────────────────────────────
ws2 = wb.create_sheet("IRB Stress Test")
ws2.freeze_panes = "A3"
ws2.merge_cells("A1:P1")
ws2["A1"].value = "ICAAP IRB STRESS TEST — Capital Impact by Scenario"
ws2["A1"].font  = Font(name="Arial", size=13, bold=True, color=WHITE)
ws2["A1"].fill  = fill(DARK_BLUE)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 28

apply_header_row(ws2, 2, ["No.","Exposure Name","Sector","IRB Asset Class","EAD",
    "Baseline PD","Baseline RW%","Baseline RWA",
    "Adverse PD","Adverse RW%","Adverse RWA",
    "Sev.Adverse PD","Sev.Adverse RW%","Sev.Adverse RWA",
    "RWA Change (Adv)","RWA Change (Sev.Adv)"])
ws2.row_dimensions[2].height = 40
for c, w in enumerate([5,28,14,20,14,12,12,14,12,12,14,14,14,14,14,14],1):
    ws2.column_dimensions[get_column_letter(c)].width = w

st_nf = [None,None,None,None,'#,##0','0.000%','0.00%','#,##0','0.000%','0.00%','#,##0','0.000%','0.00%','#,##0','#,##0;(#,##0)','#,##0;(#,##0)']
for i, loan in enumerate(portfolio):
    r   = i + 3
    bg  = LIGHT_GREY if i % 2 == 0 else WHITE
    rwa_b  = loan["irb_rwa_baseline"]
    rwa_a  = loan["irb_rwa_Adverse"]
    rwa_sa = loan["irb_rwa_Severely Adverse"]
    apply_data_row(ws2, r, [
        i+1, loan["name"], loan["sector"], loan["irb_class"], loan["ead"],
        loan["pd"], loan["irb_rw_baseline"]/100, rwa_b,
        loan["stressed_pd_Adverse"], loan["irb_rw_Adverse"]/100, rwa_a,
        loan["stressed_pd_Severely Adverse"], loan["irb_rw_Severely Adverse"]/100, rwa_sa,
        rwa_a - rwa_b, rwa_sa - rwa_b,
    ], bg=bg, number_formats=st_nf)

r_sum = len(portfolio) + 5
ws2.merge_cells(f"A{r_sum}:P{r_sum}")
ws2[f"A{r_sum}"].value = "SCENARIO SUMMARY"
ws2[f"A{r_sum}"].font  = Font(name="Arial", size=11, bold=True, color=WHITE)
ws2[f"A{r_sum}"].fill  = fill(DARK_BLUE)
ws2[f"A{r_sum}"].alignment = center()
r_sum += 1
apply_header_row(ws2, r_sum, ["Scenario","Total EAD","Total RWA","Avg RW%","Min Capital (8%)","RWA vs Baseline","Capital vs Baseline"], bg=MID_BLUE)
base_rwa = sum(l["irb_rwa_baseline"] for l in portfolio)
for scenario, mult in scenarios.items():
    r_sum += 1
    total_rwa = sum(l[f"irb_rwa_{scenario}"] for l in portfolio)
    apply_data_row(ws2, r_sum, [scenario, total_ead, total_rwa, total_rwa/total_ead,
        total_rwa*0.08, total_rwa-base_rwa, (total_rwa-base_rwa)*0.08],
        bg=YELLOW, number_formats=[None,'#,##0','#,##0','0.00%','#,##0','#,##0;(#,##0)','#,##0;(#,##0)'])

# ── SHEET 3: JURISDICTION COMPARISON ──────────────────────────────────────
ws3 = wb.create_sheet("Jurisdiction Comparison")
ws3.freeze_panes = "A3"
ws3.merge_cells("A1:P1")
ws3["A1"].value = "JURISDICTION RWA COMPARISON — Standardised Approach (Baseline)"
ws3["A1"].font  = Font(name="Arial", size=13, bold=True, color=WHITE)
ws3["A1"].fill  = fill(DARK_BLUE)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 28

apply_header_row(ws3, 2, ["No.","Exposure Name","Sector","Country","SA Exposure Class","Rating","EAD",
    "BCBS RW%","BCBS RWA","EU RW%","EU RWA","UK RW%","UK RWA","INDIA RW%","INDIA RWA","Lowest RW Jurisdiction"])
ws3.row_dimensions[2].height = 40
for c, w in enumerate([5,28,14,12,22,8,14,10,14,10,14,10,14,10,14,18],1):
    ws3.column_dimensions[get_column_letter(c)].width = w

jur_nf = [None,None,None,None,None,None,'#,##0','0.00%','#,##0','0.00%','#,##0','0.00%','#,##0','0.00%','#,##0',None]
for i, loan in enumerate(portfolio):
    r   = i + 3
    bg  = LIGHT_GREY if i % 2 == 0 else WHITE
    rws = {j.name: loan[f"sa_rw_{j.name}"] for j in jurisdictions}
    apply_data_row(ws3, r, [
        i+1, loan["name"], loan["sector"], loan["country"], str(loan["sa_class"].name),
        loan["rating"], loan["ead"],
        loan["sa_rw_BCBS"]/100,  loan["sa_rwa_BCBS"],
        loan["sa_rw_EU"]/100,    loan["sa_rwa_EU"],
        loan["sa_rw_UK"]/100,    loan["sa_rwa_UK"],
        loan["sa_rw_INDIA"]/100, loan["sa_rwa_INDIA"],
        min(rws, key=rws.get),
    ], bg=bg, number_formats=jur_nf)

r_jt = len(portfolio) + 5
ws3.merge_cells(f"A{r_jt}:P{r_jt}")
ws3[f"A{r_jt}"].value = "JURISDICTION TOTALS"
ws3[f"A{r_jt}"].font  = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3[f"A{r_jt}"].fill  = fill(DARK_BLUE)
ws3[f"A{r_jt}"].alignment = center()
r_jt += 1
apply_header_row(ws3, r_jt, ["Jurisdiction","Total EAD","Total RWA","Avg RW%","Capital (8%)","vs BCBS"], bg=MID_BLUE)
bcbs_rwa = sum(l["sa_rwa_BCBS"] for l in portfolio)
for j in jurisdictions:
    r_jt += 1
    total_rwa = sum(l[f"sa_rwa_{j.name}"] for l in portfolio)
    apply_data_row(ws3, r_jt, [j.name, total_ead, total_rwa, total_rwa/total_ead,
        total_rwa*0.08, total_rwa-bcbs_rwa],
        bg=GREEN, number_formats=[None,'#,##0','#,##0','0.00%','#,##0','#,##0;(#,##0)'])

# ── SHEET 4: EXECUTIVE SUMMARY ─────────────────────────────────────────────
ws4 = wb.create_sheet("Executive Summary")
ws4.merge_cells("A1:F1")
ws4["A1"].value = "ICAAP STRESS TESTING & JURISDICTION RWA DASHBOARD — Executive Summary"
ws4["A1"].font  = Font(name="Arial", size=14, bold=True, color=WHITE)
ws4["A1"].fill  = fill(DARK_BLUE)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 32
for c, w in enumerate([30,18,4,30,18],1):
    ws4.column_dimensions[get_column_letter(c)].width = w

def sum_title(ws, row, col, text, span=2):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    cell = ws.cell(row, col, value=text)
    cell.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    cell.fill = fill(DARK_BLUE)
    cell.alignment = center()

def kv(ws, row, col, label, value, fmt=None, bg=LIGHT_GREY):
    lc = ws.cell(row, col, value=label)
    lc.font = Font(name="Arial", size=10, bold=True)
    lc.fill = fill(bg); lc.alignment = Alignment(vertical="center"); lc.border = border()
    vc = ws.cell(row, col+1, value=value)
    vc.font = Font(name="Arial", size=10, color="0000FF")
    vc.fill = fill(WHITE); vc.alignment = right(); vc.border = border()
    if fmt: vc.number_format = fmt

sum_title(ws4, 3, 1, "PORTFOLIO OVERVIEW")
kv(ws4, 4,  1, "Total Exposures",       len(portfolio))
kv(ws4, 5,  1, "Total EAD",             total_ead,  '#,##0')
kv(ws4, 6,  1, "Derivative Contracts",  sum(1 for l in portfolio if l["is_derivative"]))
kv(ws4, 7,  1, "Derivative EAD",        sum(l["ead"] for l in portfolio if l["is_derivative"]), '#,##0')
kv(ws4, 8,  1, "Derivative EAD Share",  sum(l["ead"] for l in portfolio if l["is_derivative"])/total_ead, '0.0%')

sum_title(ws4, 10, 1, "IRB STRESS TEST SUMMARY")
apply_header_row(ws4, 11, ["Scenario","Total RWA","Avg RW%","Capital (8%)","Capital Increase"], bg=MID_BLUE)
base_rwa = sum(l["irb_rwa_baseline"] for l in portfolio)
for i, (scenario, mult) in enumerate(scenarios.items()):
    r = 12 + i
    total_rwa = sum(l[f"irb_rwa_{scenario}"] for l in portfolio)
    apply_data_row(ws4, r, [scenario, total_rwa, total_rwa/total_ead, total_rwa*0.08, (total_rwa-base_rwa)*0.08],
        bg=YELLOW if scenario != "Baseline" else WHITE,
        number_formats=[None,'#,##0','0.00%','#,##0','#,##0;(#,##0)'])

sum_title(ws4, 16, 1, "JURISDICTION COMPARISON (SA)")
apply_header_row(ws4, 17, ["Jurisdiction","Total RWA","Avg RW%","Capital (8%)","vs BCBS"], bg=MID_BLUE)
for i, j in enumerate(jurisdictions):
    r = 18 + i
    total_rwa = sum(l[f"sa_rwa_{j.name}"] for l in portfolio)
    apply_data_row(ws4, r, [j.name, total_rwa, total_rwa/total_ead, total_rwa*0.08,
        total_rwa-sum(l["sa_rwa_BCBS"] for l in portfolio)],
        bg=GREEN, number_formats=[None,'#,##0','0.00%','#,##0','#,##0;(#,##0)'])

sum_title(ws4, 23, 1, "SECTOR BREAKDOWN (IRB Baseline)")
apply_header_row(ws4, 24, ["Sector","EAD","EAD Share","RWA","Avg RW%"], bg=MID_BLUE)
sectors = {}
for loan in portfolio:
    s = loan["sector"]
    if s not in sectors: sectors[s] = {"ead": 0, "rwa": 0}
    sectors[s]["ead"] += loan["ead"]
    sectors[s]["rwa"] += loan["irb_rwa_baseline"]
for i, (sec, vals) in enumerate(sorted(sectors.items(), key=lambda x: -x[1]["ead"])):
    r = 25 + i
    apply_data_row(ws4, r, [sec, vals["ead"], vals["ead"]/total_ead, vals["rwa"],
        vals["rwa"]/vals["ead"] if vals["ead"] > 0 else 0],
        bg=LIGHT_GREY if i%2==0 else WHITE,
        number_formats=[None,'#,##0','0.0%','#,##0','0.00%'])

# ── SAVE ────────────────────────────────────────────────────────────────────
import os
output_path = "ICAAP_Dashboard.xlsx"
wb.save(output_path)
print(f"\nDashboard saved to: {output_path}")
print(f"Total exposures : {len(portfolio)}")
print(f"Total EAD       : {total_ead:,.0f}")
print(f"Sheets created  : Portfolio Detail | IRB Stress Test | Jurisdiction Comparison | Executive Summary")
import json

def convert(obj):
    if hasattr(obj, "name"):
        return obj.name
    return str(obj)

with open("portfolio.json", "w") as f:
    json.dump(portfolio, f, indent=2, default=convert)
